import csv
import os
import sys
import json
import uuid
import asyncio
import logging
import threading
import re
import contextlib
from pathlib import Path
from datetime import datetime, timezone, timedelta
from typing import List, Optional, AsyncGenerator

try:
    from googletrans import Translator
except ImportError:
    Translator = None

import base64
from array import array

import httpx
from dotenv import load_dotenv
from fastapi import FastAPI, APIRouter, HTTPException, Request, Depends, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from jose import JWTError, jwt
from passlib.context import CryptContext
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, ConfigDict

ROOT_DIR = Path(__file__).parent
sys.path.insert(0, str(ROOT_DIR.parent))
load_dotenv(ROOT_DIR / ".env", override=False)

from chatbot.shortCircuit import get_short_circuit_response

MONGO_URL = "mongodb://localhost:27017"
DB_NAME = "mizan"
MONGO_SERVER_SELECTION_TIMEOUT_MS = 5000
JWT_SECRET = "change-me"
JWT_ALGORITHM = "HS256"
JWT_ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24 * 7

USE_FINETUNED = False
FINETUNED_MODEL_PATH = "C:/tmp/mizan-chatbot/mizan-q4_k_m.gguf"
FINETUNED_HF_REPO = "olaasm/mizan"
FINETUNED_HF_FILE = "llama-3-8b-instruct.Q4_K_M.gguf"
FINETUNED_N_CTX = 2048
FINETUNED_N_THREADS = 4

CHATBOT_LOCAL_URL = ""
GEMINI_API_KEY = ""
GEMINI_MODEL = "gemini-2.5-flash"
GEMINI_MODEL_CANDIDATES = []
LAW_DATASET_PATH = "../law_dataset.csv"
LAW_DATASET_URL = ""
RAG_TOP_K = 5
RAG_MAX_CONTEXT_CHARS = 6000
GEMINI_TEMPERATURE = 0.2
GEMINI_MAX_OUTPUT_TOKENS = 512
GEMINI_REQUEST_MAX_RETRIES = 2
GEMINI_RETRY_BACKOFF_BASE = 1.5

AZURE_ML_ENDPOINT = ""
AZURE_ML_API_KEY = ""
AZURE_ML_DEPLOYMENT = ""
USE_AZURE_ENDPOINT = True
AZURE_TEMPERATURE = 0.4
AZURE_MAX_TOKENS = 128
AZURE_FREQUENCY_PENALTY = 1.15
AZURE_PRESENCE_PENALTY = 1.0
AZURE_INCLUDE_SYSTEM_PROMPT = False
AZURE_HISTORY_MAX_MESSAGES = 6
AZURE_CONTEXT_WINDOW = 512
AZURE_AUTO_CONTINUE_ROUNDS = 2
MODEL_HTTP_TIMEOUT_SECONDS = 60
STREAM_KEEPALIVE_SECONDS = 10
AZURE_REQUEST_MAX_RETRIES = 2
AZURE_RETRY_BACKOFF_BASE = 1.5


def _refresh_settings_from_env() -> None:
    global MONGO_URL, DB_NAME, MONGO_SERVER_SELECTION_TIMEOUT_MS, JWT_SECRET, JWT_ALGORITHM
    global JWT_ACCESS_TOKEN_EXPIRE_MINUTES, USE_FINETUNED, FINETUNED_MODEL_PATH, FINETUNED_HF_REPO
    global FINETUNED_HF_FILE, FINETUNED_N_CTX, FINETUNED_N_THREADS, CHATBOT_LOCAL_URL, GEMINI_API_KEY
    global GEMINI_MODEL, GEMINI_MODEL_CANDIDATES, LAW_DATASET_PATH, LAW_DATASET_URL, RAG_TOP_K, RAG_MAX_CONTEXT_CHARS
    global GEMINI_TEMPERATURE, GEMINI_MAX_OUTPUT_TOKENS, GEMINI_REQUEST_MAX_RETRIES, GEMINI_RETRY_BACKOFF_BASE
    global AZURE_ML_ENDPOINT, AZURE_ML_API_KEY, AZURE_ML_DEPLOYMENT, USE_AZURE_ENDPOINT, AZURE_TEMPERATURE
    global AZURE_MAX_TOKENS, AZURE_FREQUENCY_PENALTY, AZURE_PRESENCE_PENALTY, AZURE_INCLUDE_SYSTEM_PROMPT
    global AZURE_HISTORY_MAX_MESSAGES, AZURE_CONTEXT_WINDOW, AZURE_AUTO_CONTINUE_ROUNDS, MODEL_HTTP_TIMEOUT_SECONDS
    global STREAM_KEEPALIVE_SECONDS, AZURE_REQUEST_MAX_RETRIES, AZURE_RETRY_BACKOFF_BASE

    load_dotenv(ROOT_DIR / ".env", override=True)

    MONGO_URL = os.getenv("MONGO_URL", MONGO_URL).strip()
    DB_NAME = os.getenv("DB_NAME", DB_NAME).strip()
    MONGO_SERVER_SELECTION_TIMEOUT_MS = int(os.getenv("MONGO_SERVER_SELECTION_TIMEOUT_MS", str(MONGO_SERVER_SELECTION_TIMEOUT_MS)))
    JWT_SECRET = os.getenv("JWT_SECRET", JWT_SECRET).strip()
    JWT_ALGORITHM = os.getenv("JWT_ALGORITHM", JWT_ALGORITHM).strip()
    JWT_ACCESS_TOKEN_EXPIRE_MINUTES = int(os.getenv("JWT_ACCESS_TOKEN_EXPIRE_MINUTES", str(JWT_ACCESS_TOKEN_EXPIRE_MINUTES)))

    USE_FINETUNED = os.getenv("USE_FINETUNED", "false").lower() == "true"
    FINETUNED_MODEL_PATH = os.getenv("FINETUNED_MODEL_PATH", FINETUNED_MODEL_PATH).strip()
    FINETUNED_HF_REPO = os.getenv("FINETUNED_HF_REPO", FINETUNED_HF_REPO).strip()
    FINETUNED_HF_FILE = os.getenv("FINETUNED_HF_FILE", FINETUNED_HF_FILE).strip()
    FINETUNED_N_CTX = int(os.getenv("FINETUNED_N_CTX", str(FINETUNED_N_CTX)))
    FINETUNED_N_THREADS = int(os.getenv("FINETUNED_N_THREADS", str(FINETUNED_N_THREADS)))

    CHATBOT_LOCAL_URL = os.getenv("CHATBOT_LOCAL_URL", CHATBOT_LOCAL_URL).strip().strip('"').strip("'")

    GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", GEMINI_API_KEY).strip().strip('"').strip("'")
    GEMINI_MODEL = os.getenv("GEMINI_MODEL", GEMINI_MODEL).strip()
    GEMINI_MODEL_CANDIDATES = list(
        dict.fromkeys(  # dedupe while preserving order — duplicates double the retry burn
            candidate.strip()
            for candidate in os.getenv(
                "GEMINI_MODEL_CANDIDATES",
                ",".join(
                    [
                        GEMINI_MODEL,
                        # The -latest aliases track whatever Google currently
                        # serves, so the list survives model renames/retirements.
                        # Each model family has its own free-tier quota, so the
                        # chain keeps answers flowing when one quota runs out.
                        "gemini-flash-latest",
                        "gemini-3-flash-preview",
                        "gemini-flash-lite-latest",
                        "gemini-3.1-flash-lite",
                    ]
                ),
            ).split(",")
            if candidate.strip()
        )
    )
    LAW_DATASET_PATH = os.getenv("LAW_DATASET_PATH", LAW_DATASET_PATH).strip()
    LAW_DATASET_URL = os.getenv("LAW_DATASET_URL", LAW_DATASET_URL).strip().strip('"').strip("'")
    RAG_TOP_K = int(os.getenv("RAG_TOP_K", str(RAG_TOP_K)))
    RAG_MAX_CONTEXT_CHARS = int(os.getenv("RAG_MAX_CONTEXT_CHARS", str(RAG_MAX_CONTEXT_CHARS)))
    GEMINI_TEMPERATURE = float(os.getenv("GEMINI_TEMPERATURE", str(GEMINI_TEMPERATURE)))
    GEMINI_MAX_OUTPUT_TOKENS = int(os.getenv("GEMINI_MAX_OUTPUT_TOKENS", str(GEMINI_MAX_OUTPUT_TOKENS)))
    GEMINI_REQUEST_MAX_RETRIES = int(os.getenv("GEMINI_REQUEST_MAX_RETRIES", str(GEMINI_REQUEST_MAX_RETRIES)))
    GEMINI_RETRY_BACKOFF_BASE = float(os.getenv("GEMINI_RETRY_BACKOFF_BASE", str(GEMINI_RETRY_BACKOFF_BASE)))

    AZURE_ML_ENDPOINT = os.getenv("AZURE_ML_ENDPOINT", AZURE_ML_ENDPOINT).strip().strip('"').strip("'")
    AZURE_ML_API_KEY = os.getenv("AZURE_ML_API_KEY", AZURE_ML_API_KEY).strip().strip('"').strip("'")
    AZURE_ML_DEPLOYMENT = os.getenv("AZURE_ML_DEPLOYMENT", AZURE_ML_DEPLOYMENT).strip().strip('"').strip("'")
    USE_AZURE_ENDPOINT = os.getenv("USE_AZURE_ENDPOINT", "true").lower() == "true"
    AZURE_TEMPERATURE = float(os.getenv("AZURE_TEMPERATURE", str(AZURE_TEMPERATURE)))
    AZURE_MAX_TOKENS = int(os.getenv("AZURE_MAX_TOKENS", str(AZURE_MAX_TOKENS)))
    AZURE_FREQUENCY_PENALTY = float(os.getenv("AZURE_FREQUENCY_PENALTY", str(AZURE_FREQUENCY_PENALTY)))
    AZURE_PRESENCE_PENALTY = float(os.getenv("AZURE_PRESENCE_PENALTY", str(AZURE_PRESENCE_PENALTY)))
    AZURE_INCLUDE_SYSTEM_PROMPT = os.getenv("AZURE_INCLUDE_SYSTEM_PROMPT", "false").lower() == "true"
    AZURE_HISTORY_MAX_MESSAGES = int(os.getenv("AZURE_HISTORY_MAX_MESSAGES", str(AZURE_HISTORY_MAX_MESSAGES)))
    AZURE_CONTEXT_WINDOW = int(os.getenv("AZURE_CONTEXT_WINDOW", str(AZURE_CONTEXT_WINDOW)))
    AZURE_AUTO_CONTINUE_ROUNDS = int(os.getenv("AZURE_AUTO_CONTINUE_ROUNDS", str(AZURE_AUTO_CONTINUE_ROUNDS)))
    MODEL_HTTP_TIMEOUT_SECONDS = float(os.getenv("MODEL_HTTP_TIMEOUT_SECONDS", str(MODEL_HTTP_TIMEOUT_SECONDS)))
    STREAM_KEEPALIVE_SECONDS = float(os.getenv("STREAM_KEEPALIVE_SECONDS", str(STREAM_KEEPALIVE_SECONDS)))
    AZURE_REQUEST_MAX_RETRIES = int(os.getenv("AZURE_REQUEST_MAX_RETRIES", str(AZURE_REQUEST_MAX_RETRIES)))
    AZURE_RETRY_BACKOFF_BASE = float(os.getenv("AZURE_RETRY_BACKOFF_BASE", str(AZURE_RETRY_BACKOFF_BASE)))


_refresh_settings_from_env()

_llm_instance = None
_llm_lock = asyncio.Lock()
_law_dataset_cache: Optional[List[dict]] = None
_law_token_index: dict = {}
_law_dataset_lock = threading.Lock()

client = AsyncIOMotorClient(
    MONGO_URL,
    serverSelectionTimeoutMS=MONGO_SERVER_SELECTION_TIMEOUT_MS,
)
db = client[DB_NAME]

# Configure logging to output to stderr with proper formatting
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    force=True,
)
logger = logging.getLogger("mizan")
logger.setLevel(logging.INFO)
logger.propagate = True
if not logger.handlers:
    handler = logging.StreamHandler()
    handler.setLevel(logging.INFO)
    formatter = logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s")
    handler.setFormatter(formatter)
    logger.addHandler(handler)

logger.info("Backend logging initialized; gemini_configured=%s", bool(GEMINI_API_KEY))


def _log_answer_source(source: str, chat_id: str, user_id: str, reply_text: Optional[str]) -> None:
    preview = (reply_text or "").strip().replace("\n", " ")[:120]
    logger.info("Answer source=%s chat_id=%s user=%s preview=%s", source, chat_id, user_id, preview or "<empty>")


def _log_message_event(chat_id: str, user_id: str, role: str, origin: str, message_text: Optional[str]) -> None:
    preview = (message_text or "").strip().replace("\n", " ")[:120]
    logger.info("MessageEvent chat_id=%s role=%s origin=%s user=%s preview=%s", chat_id, role, origin, user_id, preview or "<empty>")

# Initialize translator for Arabic translation (lazy-loaded)
translator = None

async def _get_translator():
    """Get or initialize the translator"""
    global translator
    if translator is None and Translator:
        translator = Translator()
    return translator

app = FastAPI(title="Mizan API")
api = APIRouter(prefix="/api")


@app.middleware("http")
async def refresh_settings_middleware(request: Request, call_next):
    _refresh_settings_from_env()
    logger.info("Incoming %s %s | gemini_configured=%s", request.method, request.url.path, bool(GEMINI_API_KEY))
    response = await call_next(request)
    return response

SYSTEM_PROMPT = """You are Mizan, a Lebanese legal assistant specialized in Lebanese law. Answer all questions based exclusively on Lebanese laws. Cite article numbers. You will receive questions in both Arabic and English. Always respond in Arabic."""


class Chat(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str = "محادثة جديدة"
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class Attachment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    filename: str
    mime_type: str
    size: int


class Message(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    chat_id: str
    role: str
    content: str
    source: Optional[str] = None
    attachment: Optional[Attachment] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()


class UserCreate(BaseModel):
    model_config = ConfigDict(extra="ignore")
    name: str
    email: str
    password: str


class UserLogin(BaseModel):
    model_config = ConfigDict(extra="ignore")
    email: str
    password: str


class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    email: str
    hashed_password: str
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class UserOut(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    email: str
    created_at: str


class TokenResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    token: str


def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)


def hash_password(password: str) -> str:
    if len(password.encode("utf-8")) > 72:
        raise ValueError("كلمة المرور يجب أن لا تتجاوز 72 بايت")
    return pwd_context.hash(password)


def create_access_token(subject_id: str) -> str:
    expire = datetime.now(timezone.utc) + timedelta(minutes=JWT_ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode = {"sub": subject_id, "exp": expire}
    return jwt.encode(to_encode, JWT_SECRET, algorithm=JWT_ALGORITHM)


async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> UserOut:
    token = credentials.credentials
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("sub")
        if not user_id:
            raise ValueError("Missing user id in token")
    except (JWTError, ValueError):
        raise HTTPException(401, "Invalid or expired authentication token")

    user_doc = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user_doc:
        raise HTTPException(401, "User not found")
    return UserOut(**user_doc)


async def _create_guest_user() -> UserOut:
    """Create a fresh, isolated guest account per guest session.

    Each visitor gets their own user so guests can never see each other's
    chats. The random email/password are never shown; the session lives in
    the issued JWT only.
    """
    guest_user = User(
        name="Guest",
        email=f"guest-{uuid.uuid4().hex}@mizan.local",
        hashed_password=hash_password(f"guest-{uuid.uuid4().hex}"),
    )
    doc = guest_user.model_dump()
    doc["is_guest"] = True  # marker for future cleanup of stale guest accounts
    await db.users.insert_one(doc)
    return UserOut(**guest_user.model_dump())


class ChatCreate(BaseModel):
    title: Optional[str] = None


class ChatUpdate(BaseModel):
    title: str


class SendMessageBody(BaseModel):
    chat_id: str
    content: str
    use_local: bool = False
    use_azure: bool = False
    use_finetuned: bool = False


async def _generate_title_from_first_message(text: str) -> str:
    t = text.strip().replace("\n", " ")
    return (t[:40] + "…") if len(t) > 40 else (t or "محادثة جديدة")


async def _call_local_chatbot(history: List[dict]) -> Optional[str]:
    if not CHATBOT_LOCAL_URL:
        return None
    try:
        async with httpx.AsyncClient(timeout=MODEL_HTTP_TIMEOUT_SECONDS) as cx:
            r = await cx.post(
                f"{CHATBOT_LOCAL_URL.rstrip('/')}/chat",
                json={"system": SYSTEM_PROMPT, "messages": history},
            )
            r.raise_for_status()
            data = r.json()
            return data.get("reply")
    except Exception as e:
        logger.warning("Local chatbot unreachable: %s", e)
        return None


def _resolve_law_dataset_path() -> Path:
    path = Path(LAW_DATASET_PATH)
    if not path.is_absolute():
        path = (ROOT_DIR / path).resolve()
    return path


def _normalize_rag_text(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").lower()).strip()


def _tokenize_rag_text(text: str) -> List[str]:
    return re.findall(r"[\u0600-\u06FF\w]+", _normalize_rag_text(text))


def _convert_xlsx_to_csv(xlsx_path: Path, csv_path: Path) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        with csv_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            for row in worksheet.iter_rows(values_only=True):
                writer.writerow(["" if cell is None else cell for cell in row])
    finally:
        workbook.close()


def _download_law_dataset(dataset_path: Path) -> bool:
    """Fetch the private law dataset from LAW_DATASET_URL (used in deployments
    where the spreadsheet is intentionally kept out of the public git repo).

    Accepts either a CSV or an XLSX workbook; XLSX is converted to the CSV
    the loader expects. Dropbox share links are normalized to direct downloads.
    """
    if not LAW_DATASET_URL:
        return False
    try:
        url = LAW_DATASET_URL
        if "dropbox.com" in url:
            url = url.replace("dl=0", "dl=1")
        logger.info("Law dataset missing; downloading from LAW_DATASET_URL ...")
        dataset_path.parent.mkdir(parents=True, exist_ok=True)
        tmp_path = dataset_path.with_name(dataset_path.name + ".part")
        with httpx.Client(timeout=300, follow_redirects=True) as cx:
            with cx.stream("GET", url) as response:
                response.raise_for_status()
                with tmp_path.open("wb") as handle:
                    for chunk in response.iter_bytes():
                        handle.write(chunk)

        with tmp_path.open("rb") as handle:
            magic = handle.read(4)

        if magic[:2] == b"PK":
            # XLSX workbook (zip container) — convert to CSV for the loader
            logger.info("Downloaded dataset is an XLSX workbook; converting to CSV ...")
            xlsx_path = tmp_path.with_name(tmp_path.name + ".xlsx")  # openpyxl requires the extension
            tmp_path.replace(xlsx_path)
            try:
                _convert_xlsx_to_csv(xlsx_path, dataset_path)
            finally:
                xlsx_path.unlink()
        elif magic[:1] == b"<":
            # HTML page (e.g. a Dropbox preview link) — not the file itself
            tmp_path.unlink()
            raise RuntimeError(
                "LAW_DATASET_URL returned an HTML page, not the dataset. "
                "For Dropbox links make sure the URL ends with dl=1."
            )
        else:
            tmp_path.replace(dataset_path)

        logger.info("Law dataset ready at %s (%d bytes)", dataset_path, dataset_path.stat().st_size)
        return True
    except Exception as e:
        logger.exception("Failed to download law dataset: %s", e)
        return False


def _load_law_dataset() -> List[dict]:
    global _law_dataset_cache, _law_token_index
    if _law_dataset_cache is not None:
        return _law_dataset_cache

    with _law_dataset_lock:
        if _law_dataset_cache is not None:
            return _law_dataset_cache

        dataset_path = _resolve_law_dataset_path()
        if not dataset_path.exists():
            _download_law_dataset(dataset_path)
        if not dataset_path.exists():
            logger.warning("Law dataset not found at %s", dataset_path)
            _law_dataset_cache = []
            return _law_dataset_cache

        # Static per-record score components (type prior + boosted phrases) are
        # precomputed at load time; the token overlap uses a global inverted
        # index instead of a per-record token set. This keeps scoring identical
        # to the old per-record version while fitting in ~1/3 of the RAM —
        # required for 512MB hosts like Render's free tier.
        boosted_phrases = (
            "قانون العمل",
            "الصرف من الخدمة",
            "تعويض الصرف",
            "تعويض الصرف من الخدمة",
            "الصرف التعسفي",
            "فصل تعسفي",
        )
        row_fields = ("nationality", "questions", "Instruction", "output_text", "type", "input_text")

        records: List[dict] = []
        token_index: dict = {}
        try:
            with dataset_path.open("r", encoding="utf-8-sig", newline="") as handle:
                reader = csv.DictReader(handle)
                for index, row in enumerate(reader, start=1):
                    text_parts = [
                        str(row.get("questions", "") or "").strip(),
                        str(row.get("Instruction", "") or "").strip(),
                        str(row.get("output_text", "") or "").strip(),
                        str(row.get("input_text", "") or "").strip(),
                    ]
                    combined_text = "\n".join(part for part in text_parts if part)
                    if not combined_text:
                        continue

                    normalized_text = _normalize_rag_text(combined_text)
                    record_type = _normalize_rag_text(str(row.get("type", "") or ""))

                    static_bonus = 0
                    if record_type == "law":
                        static_bonus += 18
                    elif record_type == "case":
                        static_bonus -= 8
                    for phrase in boosted_phrases:
                        if phrase in normalized_text:
                            static_bonus += 8

                    position = len(records)
                    for token in set(_tokenize_rag_text(combined_text)):
                        token_index.setdefault(sys.intern(token), array("I")).append(position)

                    records.append(
                        {
                            "index": index,
                            "row": {field: str(row.get(field, "") or "") for field in row_fields},
                            "type": record_type,
                            "text_len": len(combined_text),
                            "normalized": normalized_text,
                            "static_bonus": static_bonus,
                            "match_fields": tuple(
                                _normalize_rag_text(str(row.get(field, "") or ""))
                                for field in ("questions", "Instruction", "type", "nationality")
                            ),
                        }
                    )
        except Exception as e:
            logger.exception("Failed to load law dataset from %s: %s", dataset_path, e)
            records = []
            token_index = {}

        _law_token_index = token_index
        _law_dataset_cache = records
        logger.info("Loaded %d law dataset rows from %s", len(records), dataset_path)
        return _law_dataset_cache


def _retrieve_law_context(query_text: str, top_k: int) -> List[dict]:
    records = _load_law_dataset()
    if not records:
        return []

    normalized_query = _normalize_rag_text(query_text)
    query_tokens = set(_tokenize_rag_text(normalized_query))

    # Token overlap per record via the inverted index — equal to
    # len(query_tokens & record_tokens) in the old per-record-set version.
    overlap_counts: dict = {}
    for token in query_tokens:
        for position in _law_token_index.get(token, ()):
            overlap_counts[position] = overlap_counts.get(position, 0) + 1

    scored_records = []
    for position, record in enumerate(records):
        score = overlap_counts.get(position, 0) + record["static_bonus"]
        if normalized_query and normalized_query in record["normalized"]:
            score += 12
        for value in record["match_fields"]:
            if value and value in normalized_query:
                score += 3
        scored_records.append((score, record))

    scored_records.sort(key=lambda item: (item[0], item[1]["text_len"]), reverse=True)

    selected = [record for score, record in scored_records if score > 0]
    law_selected = [record for record in selected if record.get("type") == "law"]

    if law_selected:
        return law_selected[:top_k]
    if selected:
        return selected[:top_k]
    return [record for _, record in scored_records[:top_k]]


def _format_rag_context(records: List[dict]) -> str:
    if not records:
        return ""

    sections: List[str] = []
    total_chars = 0

    for idx, record in enumerate(records, start=1):
        row = record["row"]
        header = str(row.get("Instruction", "") or row.get("type", "") or f"Document {record['index']}").strip()
        details = [f"[{idx}] {header}"]

        for field in ("nationality", "questions", "type"):
            value = str(row.get(field, "") or "").strip()
            if value:
                details.append(f"{field}: {value}")

        input_text = str(row.get("input_text", "") or "").strip()
        output_text = str(row.get("output_text", "") or "").strip()
        if input_text:
            details.append(f"law_text: {input_text}")
        if output_text:
            details.append(f"reference_answer: {output_text}")

        block = "\n".join(details).strip()
        if not block:
            continue

        if total_chars + len(block) > RAG_MAX_CONTEXT_CHARS:
            remaining = max(0, RAG_MAX_CONTEXT_CHARS - total_chars)
            if remaining <= 0:
                break
            block = block[:remaining].rstrip()

        sections.append(block)
        total_chars += len(block)

        if total_chars >= RAG_MAX_CONTEXT_CHARS:
            break

    return "\n\n".join(sections)


def _build_gemini_prompt(history: List[dict], context: str) -> str:
    recent_history = history[-6:]
    history_lines = []
    for message in recent_history:
        role = message.get("role", "user")
        label = "User" if role == "user" else "Assistant"
        content = str(message.get("content", "") or "").strip()
        if content:
            history_lines.append(f"{label}: {content}")

    history_text = "\n".join(history_lines).strip()
    context_text = context.strip() if context.strip() else "No closely matching passages were found in the dataset."

    return (
        "Use the retrieved Lebanese law context below when it is relevant to the user's question. "
        "If the context does not contain the answer, answer honestly from your own knowledge of Lebanese law instead — "
        "never reply that you could not find the answer or that the information is unavailable. "
        "When your answer relies on general knowledge rather than the retrieved context, end with a brief note "
        "(in Arabic) that the answer is based on general legal knowledge and it is best to verify with a lawyer.\n\n"
        f"Retrieved context:\n{context_text}\n\n"
        f"Recent conversation:\n{history_text or 'No prior conversation.'}\n\n"
        "Answer the latest user question in Arabic and cite article numbers whenever you know them."
    )


GEMINI_QUOTA_MESSAGE = (
    "وصلت الخدمة إلى الحد الأقصى للاستخدام المجاني اليوم. "
    "يرجى المحاولة مجدداً بعد ساعات قليلة."
)


def _friendly_gemini_error(technical_message: str) -> str:
    """Map quota-exhaustion errors to a user-facing Arabic message; the full
    technical detail still goes to the server log for debugging."""
    lowered = technical_message.lower()
    if "429" in technical_message or "quota" in lowered or "resource_exhausted" in lowered:
        logger.warning("Gemini quota exhausted: %s", technical_message[:300])
        return GEMINI_QUOTA_MESSAGE
    return technical_message


def _extract_reply_from_gemini(data: dict) -> Optional[str]:
    candidates = data.get("candidates") or []
    for candidate in candidates:
        content = candidate.get("content") or {}
        parts = content.get("parts") or []
        text = "".join(part.get("text", "") for part in parts if isinstance(part, dict))
        if isinstance(text, str) and text.strip():
            return text.strip()
    return None


async def _call_gemini_rag(history: List[dict]) -> Optional[str]:
    if not GEMINI_API_KEY:
        return None

    query_text = "\n".join(str(message.get("content", "") or "") for message in history[-6:]).strip()
    # Retrieval (and the first-call dataset download/parse) is CPU/IO heavy —
    # run it off the event loop so slow hosts don't stall every request.
    records = await asyncio.to_thread(_retrieve_law_context, query_text, max(1, RAG_TOP_K))
    context = _format_rag_context(records)

    # If no retrieved context is available, fall back to a generative prompt
    # so the model can answer from its general knowledge instead of replying
    # that nothing was found in the dataset.
    generative_mode = not bool(context.strip())

    if generative_mode:
        # Use recent user messages as the query for a generative answer.
        prompt_contents = query_text or (history[-1].get("content", "") if history else "")
        payload = {
            "systemInstruction": {"parts": [{"text": SYSTEM_PROMPT}]},
            "contents": [{"role": "user", "parts": [{"text": prompt_contents}]}],
            "generationConfig": {
                "temperature": GEMINI_TEMPERATURE,
                "maxOutputTokens": GEMINI_MAX_OUTPUT_TOKENS,
            },
        }
    else:
        prompt = _build_gemini_prompt(history, context)
        payload = {
            "systemInstruction": {"parts": [{"text": SYSTEM_PROMPT}]},
            "contents": [{"role": "user", "parts": [{"text": prompt}]}],
            "generationConfig": {
                "temperature": GEMINI_TEMPERATURE,
                "maxOutputTokens": GEMINI_MAX_OUTPUT_TOKENS,
            },
        }

    try:
        async with httpx.AsyncClient(timeout=MODEL_HTTP_TIMEOUT_SECONDS) as cx:
            last_error_message = ""
            saw_quota = False
            for model_name in GEMINI_MODEL_CANDIDATES:
                url = (
                    f"https://generativelanguage.googleapis.com/v1beta/models/"
                    f"{model_name}:generateContent?key={GEMINI_API_KEY}"
                )
                data = None

                try:
                    last_exc = None
                    for attempt in range(1, GEMINI_REQUEST_MAX_RETRIES + 2):
                        try:
                            response = await cx.post(url, json=payload)
                            response.raise_for_status()
                            data = response.json()
                            break
                        except httpx.HTTPStatusError as e:
                            status = getattr(e.response, "status_code", None)
                            detail = ""
                            try:
                                detail = e.response.text[:500]
                            except Exception:
                                detail = ""

                            if status == 404:
                                last_error_message = f"{model_name}: {detail or 'model not available'}"
                                logger.warning("Gemini model %s unavailable, trying next candidate", model_name)
                                break

                            if status == 429:
                                # Quota won't recover within a retry backoff —
                                # go straight to the next model (separate quota).
                                saw_quota = True
                                last_error_message = f"{model_name}: HTTP 429: {detail}"
                                logger.warning("Gemini model %s quota exhausted, trying next candidate", model_name)
                                break

                            if status in {408, 500, 502, 503, 504} and attempt <= GEMINI_REQUEST_MAX_RETRIES:
                                backoff = GEMINI_RETRY_BACKOFF_BASE ** (attempt - 1)
                                logger.warning("Gemini HTTP %s on attempt %s for %s, retrying after %.1fs", status, attempt, model_name, backoff)
                                await asyncio.sleep(backoff)
                                last_exc = e
                                continue
                            raise
                        except httpx.TimeoutException as e:
                            if attempt <= GEMINI_REQUEST_MAX_RETRIES:
                                backoff = GEMINI_RETRY_BACKOFF_BASE ** (attempt - 1)
                                logger.warning("Gemini request timed out on attempt %s for %s, retrying after %.1fs", attempt, model_name, backoff)
                                await asyncio.sleep(backoff)
                                last_exc = e
                                continue
                            raise

                    if last_exc is not None and not isinstance(last_exc, httpx.HTTPStatusError):
                        raise last_exc

                    if data is None:
                        continue

                    if isinstance(data, dict):
                        error = data.get("error")
                        if error:
                            last_error_message = f"{model_name}: {error}"
                            continue

                        reply = _extract_reply_from_gemini(data)
                        if reply:
                            return reply

                        block_reason = (data.get("promptFeedback") or {}).get("blockReason")
                        if block_reason:
                            last_error_message = f"{model_name}: blocked ({block_reason})"
                            continue

                    last_error_message = f"{model_name}: Gemini response missing expected content"
                except httpx.HTTPStatusError as e:
                    detail = ""
                    try:
                        detail = e.response.text[:500]
                    except Exception:
                        detail = ""
                    last_error_message = f"{model_name}: HTTP {e.response.status_code}: {detail}"
                    continue

        if saw_quota:
            logger.warning("All Gemini candidates exhausted (quota). Last error: %s", last_error_message[:300])
            raise RuntimeError(GEMINI_QUOTA_MESSAGE)
        raise RuntimeError(
            _friendly_gemini_error(
                "Gemini response missing expected content. Tried models: "
                f"{', '.join(GEMINI_MODEL_CANDIDATES)}. Last error: {last_error_message}"
            )
        )
    except httpx.HTTPStatusError as e:
        detail = ""
        try:
            detail = e.response.text[:500]
        except Exception:
            detail = ""
        raise RuntimeError(_friendly_gemini_error(f"Gemini HTTP {e.response.status_code}: {detail}"))
    except httpx.TimeoutException:
        raise RuntimeError("انتهت مهلة الاتصال بالمساعد. يرجى المحاولة مرة أخرى.")
    except RuntimeError:
        raise
    except Exception as e:
        raise RuntimeError(f"Gemini API unreachable: {e}")


DOCUMENT_MAX_BYTES = 15 * 1024 * 1024  # keep comfortably under Gemini's inline request size limit
DOCUMENT_ALLOWED_MIME_TYPES = {
    "application/pdf",
    "text/plain",
    "image/png",
    "image/jpeg",
}
UPLOADS_DIR = ROOT_DIR / "uploads"

PDF_FONT_CANDIDATES = [
    os.getenv("PDF_FONT_PATH", "").strip(),
    str(ROOT_DIR / "fonts" / "Amiri-Regular.ttf"),  # bundled, works on any host
    "C:/Windows/Fonts/arial.ttf",
    "C:/Windows/Fonts/tahoma.ttf",
]
PDF_BOLD_FONT_CANDIDATES = [
    os.getenv("PDF_BOLD_FONT_PATH", "").strip(),
    str(ROOT_DIR / "fonts" / "Amiri-Bold.ttf"),
    "C:/Windows/Fonts/arialbd.ttf",
    "C:/Windows/Fonts/tahomabd.ttf",
]

_PDF_REQUEST_KEYWORDS = ("pdf", "بي دي اف", "بي دي إف")


def _wants_pdf(text: str) -> bool:
    lowered = (text or "").lower()
    return any(keyword in lowered for keyword in _PDF_REQUEST_KEYWORDS)


def _resolve_pdf_font(candidates: List[str]) -> Optional[str]:
    for candidate in candidates:
        if candidate and os.path.isfile(candidate):
            return candidate
    return None


def _generate_answer_pdf_sync(text: str) -> bytes:
    from fpdf import FPDF  # imported lazily so the server runs even without fpdf2 installed

    font_path = _resolve_pdf_font(PDF_FONT_CANDIDATES)
    if not font_path:
        raise RuntimeError("No Arabic-capable TTF font found for PDF generation (set PDF_FONT_PATH)")
    bold_path = _resolve_pdf_font(PDF_BOLD_FONT_CANDIDATES) or font_path

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.add_font("body", style="", fname=font_path)
    pdf.add_font("body", style="B", fname=bold_path)
    pdf.set_text_shaping(True)
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()

    pdf.set_font("body", style="B", size=18)
    pdf.set_text_color(20, 83, 45)
    pdf.cell(0, 12, "ميزان", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("body", size=9)
    pdf.set_text_color(120, 120, 120)
    pdf.cell(
        0,
        6,
        "مساعد قانوني - المعلومات للاطلاع العام ولا تعتبر استشارة قانونية رسمية",
        align="C",
        new_x="LMARGIN",
        new_y="NEXT",
    )
    pdf.ln(4)
    pdf.set_draw_color(200, 200, 200)
    pdf.line(pdf.l_margin, pdf.get_y(), pdf.w - pdf.r_margin, pdf.get_y())
    pdf.ln(6)

    pdf.set_text_color(30, 30, 30)
    pdf.set_font("body", size=12)
    for paragraph in (text or "").split("\n"):
        if not paragraph.strip():
            pdf.ln(4)
            continue
        pdf.multi_cell(0, 8, paragraph, align="R", new_x="LMARGIN", new_y="NEXT")

    return bytes(pdf.output())


async def _generate_answer_pdf(text: str) -> bytes:
    return await asyncio.to_thread(_generate_answer_pdf_sync, text)


PDF_SYSTEM_NOTE = (
    "\n\n(ملاحظة للنظام: المستخدم طلب الإجابة كملف PDF. "
    "اكتب محتوى الإجابة كاملاً ومنسقاً كمستند، ولا تعتذر عن عدم قدرتك على إنشاء ملفات — "
    "سيقوم النظام بتحويل إجابتك تلقائياً إلى ملف PDF وإرفاقه.)"
)


def _append_pdf_note_to_history(history: List[dict]) -> List[dict]:
    """Return a copy of history with a system note on the last user message so the
    model writes document content instead of refusing to produce a file."""
    if not history:
        return history
    adjusted = [dict(message) for message in history]
    for message in reversed(adjusted):
        if message.get("role") == "user":
            message["content"] = str(message.get("content", "")) + PDF_SYSTEM_NOTE
            break
    return adjusted


async def _attach_answer_pdf(ai_msg: Message, chat_title: str) -> None:
    """Render the assistant reply to a PDF and attach it to the message (best effort)."""
    try:
        pdf_bytes = await _generate_answer_pdf(ai_msg.content)
    except Exception as e:
        logger.warning("PDF generation failed, sending text-only answer: %s", e)
        return

    filename = "mizan-answer.pdf"
    title = (chat_title or "").strip()
    if title and title != "محادثة جديدة":
        safe_title = re.sub(r"[^\w؀-ۿ-]+", "-", title).strip("-")[:40]
        if safe_title:
            filename = f"{safe_title}.pdf"

    _save_uploaded_document(ai_msg.chat_id, ai_msg.id, filename, pdf_bytes)
    ai_msg.attachment = Attachment(filename=filename, mime_type="application/pdf", size=len(pdf_bytes))


def _save_uploaded_document(chat_id: str, message_id: str, original_filename: str, file_bytes: bytes) -> Path:
    chat_dir = UPLOADS_DIR / chat_id
    chat_dir.mkdir(parents=True, exist_ok=True)
    safe_name = os.path.basename(original_filename or "document")
    dest = chat_dir / f"{message_id}_{safe_name}"
    dest.write_bytes(file_bytes)
    return dest


def _resolve_uploaded_document(chat_id: str, message_id: str) -> Optional[Path]:
    chat_dir = UPLOADS_DIR / chat_id
    if not chat_dir.is_dir():
        return None
    for candidate in chat_dir.glob(f"{message_id}_*"):
        if candidate.is_file():
            return candidate
    return None


async def _call_gemini_document(question: str, file_bytes: bytes, mime_type: str) -> Optional[str]:
    if not GEMINI_API_KEY:
        return None

    encoded = base64.standard_b64encode(file_bytes).decode("utf-8")
    payload = {
        "systemInstruction": {"parts": [{"text": SYSTEM_PROMPT}]},
        "contents": [
            {
                "role": "user",
                "parts": [
                    {"inline_data": {"mime_type": mime_type, "data": encoded}},
                    {"text": question.strip() or "Summarize this document and cite the key legal points."},
                ],
            }
        ],
        "generationConfig": {
            "temperature": GEMINI_TEMPERATURE,
            "maxOutputTokens": GEMINI_MAX_OUTPUT_TOKENS,
        },
    }

    try:
        async with httpx.AsyncClient(timeout=MODEL_HTTP_TIMEOUT_SECONDS) as cx:
            last_error_message = ""
            saw_quota = False
            for model_name in GEMINI_MODEL_CANDIDATES:
                url = (
                    f"https://generativelanguage.googleapis.com/v1beta/models/"
                    f"{model_name}:generateContent?key={GEMINI_API_KEY}"
                )
                try:
                    response = await cx.post(url, json=payload)
                    response.raise_for_status()
                    data = response.json()
                except httpx.HTTPStatusError as e:
                    status = getattr(e.response, "status_code", None)
                    detail = ""
                    try:
                        detail = e.response.text[:500]
                    except Exception:
                        detail = ""
                    if status in {404, 429}:
                        if status == 429:
                            saw_quota = True
                        last_error_message = f"{model_name}: HTTP {status}: {detail or 'model not available'}"
                        logger.warning("Gemini model %s failed (%s) for document analysis, trying next candidate", model_name, status)
                        continue
                    raise RuntimeError(_friendly_gemini_error(f"Gemini HTTP {status}: {detail}"))

                if isinstance(data, dict):
                    error = data.get("error")
                    if error:
                        last_error_message = f"{model_name}: {error}"
                        continue
                    reply = _extract_reply_from_gemini(data)
                    if reply:
                        return reply
                    block_reason = (data.get("promptFeedback") or {}).get("blockReason")
                    if block_reason:
                        last_error_message = f"{model_name}: blocked ({block_reason})"
                        continue
                    last_error_message = f"{model_name}: Gemini response missing expected content"

        if saw_quota:
            logger.warning("All Gemini candidates exhausted for document analysis (quota). Last error: %s", last_error_message[:300])
            raise RuntimeError(GEMINI_QUOTA_MESSAGE)
        raise RuntimeError(
            _friendly_gemini_error(
                "Gemini document analysis failed. Tried models: "
                f"{', '.join(GEMINI_MODEL_CANDIDATES)}. Last error: {last_error_message}"
            )
        )
    except httpx.TimeoutException:
        raise RuntimeError("انتهت مهلة الاتصال بالمساعد. يرجى المحاولة مرة أخرى.")


async def _stream_gemini_rag(history: List[dict]) -> AsyncGenerator[str, None]:
    reply = await _call_gemini_rag(history)
    if not reply:
        raise RuntimeError("Gemini API unreachable or returned empty reply")

    for part in re.findall(r"\S+\s*", reply):
        yield part
        await asyncio.sleep(0.01)


def _azure_headers() -> dict:
    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json",
        "Authorization": f"Bearer {AZURE_ML_API_KEY}",
    }
    if AZURE_ML_DEPLOYMENT:
        headers["azureml-model-deployment"] = AZURE_ML_DEPLOYMENT
    return headers


def _extract_reply_from_azure(data: dict) -> Optional[str]:
    choices = data.get("choices") or []
    if not choices:
        return None
    message = choices[0].get("message") or {}
    content = message.get("content")
    if isinstance(content, str):
        return content.strip()
    return None


def _extract_finish_reason_from_azure(data: dict) -> Optional[str]:
    choices = data.get("choices") or []
    if not choices:
        return None
    reason = choices[0].get("finish_reason")
    if isinstance(reason, str):
        return reason
    return None


def _looks_truncated(reply: str, finish_reason: Optional[str]) -> bool:
    if not reply:
        return False
    if finish_reason in {"length", "max_tokens"}:
        return True

    # If the model stops without closing punctuation, try one continuation pass.
    return bool(re.search(r"[\w\u0600-\u06FF]\s*$", reply))


def _build_azure_messages(history: List[dict]) -> List[dict]:
    """Build messages for Azure, ensuring valid conversation structure
    
    Strategy: Keep the most recent message (user's current question) and 
    one or more prior messages to provide context, prioritizing assistant messages.
    """
    if not history:
        return []
    
    logger.info("_build_azure_messages: input history=%d", len(history))

    def _approx_tokens(text: str) -> int:
        # Conservative token estimate: words ~= tokens
        if not text:
            return 0
        return max(1, len(re.findall(r"\S+", text)))

    # Reserve tokens for the model's response
    reserved_for_response = AZURE_MAX_TOKENS
    max_context = AZURE_CONTEXT_WINDOW
    available_for_context = max(0, max_context - reserved_for_response)

    msgs: List[dict] = []
    system_tokens = 0
    
    # Add system prompt if enabled
    if AZURE_INCLUDE_SYSTEM_PROMPT:
        system_tokens = _approx_tokens(SYSTEM_PROMPT) + 4
        if system_tokens >= available_for_context:
            short_sys = SYSTEM_PROMPT[:1024]
            msgs.append({"role": "system", "content": short_sys})
            available_for_context = max(0, available_for_context - _approx_tokens(short_sys) - 4)
        else:
            msgs.append({"role": "system", "content": SYSTEM_PROMPT})
            available_for_context = max(0, available_for_context - system_tokens)

    # Build conversation history: keep most recent messages that fit in token budget
    # Prioritize including: last user message + at least one recent assistant message if available
    kept: List[dict] = []
    tokens = 0
    
    # Iterate from most recent backwards
    for idx in range(len(history) - 1, -1, -1):
        msg = history[idx]
        msg_tokens = _approx_tokens(str(msg.get("content", ""))) + 4
        
        # Always keep the most recent message (current user question)
        if idx == len(history) - 1:
            kept.insert(0, msg)
            tokens += msg_tokens
            continue
        
        # For earlier messages, only keep if we have token budget
        if tokens + msg_tokens <= available_for_context:
            kept.insert(0, msg)
            tokens += msg_tokens
            # Stop after collecting enough context (max AZURE_HISTORY_MAX_MESSAGES)
            if len(kept) >= AZURE_HISTORY_MAX_MESSAGES:
                break
        else:
            break
    
    # Context is now handled by English system prompt to avoid encoding issues
    
    msgs.extend(kept)
    logger.info("_build_azure_messages: returning %d messages (tokens=%d)", len(msgs), tokens)
    return msgs


async def _call_azure_endpoint(history: List[dict]) -> Optional[str]:
    if not AZURE_ML_ENDPOINT or not AZURE_ML_API_KEY:
        return None

    msgs = _build_azure_messages(history)
    collected_parts: List[str] = []

    try:
        async with httpx.AsyncClient(timeout=MODEL_HTTP_TIMEOUT_SECONDS) as cx:
            for round_idx in range(max(0, AZURE_AUTO_CONTINUE_ROUNDS) + 1):
                payload = {
                    "messages": msgs,
                    "temperature": AZURE_TEMPERATURE,
                    "max_tokens": AZURE_MAX_TOKENS,
                    "frequency_penalty": AZURE_FREQUENCY_PENALTY,
                    "presence_penalty": AZURE_PRESENCE_PENALTY,
                }
                # Log first 500 chars of payload (truncate for readability)
                logger.info("Sending to Azure: %s", json.dumps(payload, ensure_ascii=False)[:500])

                last_exc = None
                for attempt in range(1, AZURE_REQUEST_MAX_RETRIES + 2):
                    try:
                        # Ensure UTF-8 encoding for the request body
                        r = await cx.post(
                            AZURE_ML_ENDPOINT, 
                            headers=_azure_headers(),
                            json=payload,
                            # httpx will automatically UTF-8 encode the JSON content
                        )
                        r.raise_for_status()
                        data = r.json()
                        break
                    except httpx.HTTPStatusError as e:
                        status = getattr(e.response, "status_code", None)
                        # Retry on transient server/client errors like 408/429/5xx
                        if status in {408, 429, 500, 502, 503, 504} and attempt <= AZURE_REQUEST_MAX_RETRIES:
                            backoff = AZURE_RETRY_BACKOFF_BASE ** (attempt - 1)
                            logger.warning("Azure HTTP %s on attempt %s, retrying after %.1fs", status, attempt, backoff)
                            await asyncio.sleep(backoff)
                            last_exc = e
                            continue
                        # otherwise re-raise
                        raise
                    except httpx.TimeoutException as e:
                        if attempt <= AZURE_REQUEST_MAX_RETRIES:
                            backoff = AZURE_RETRY_BACKOFF_BASE ** (attempt - 1)
                            logger.warning("Azure request timed out on attempt %s, retrying after %.1fs", attempt, backoff)
                            await asyncio.sleep(backoff)
                            last_exc = e
                            continue
                        raise

                if last_exc is not None and (not isinstance(last_exc, httpx.HTTPStatusError)):
                    # propagate last exception if retries exhausted
                    raise last_exc

                if isinstance(data, dict) and data.get("error"):
                    raise RuntimeError(f"Azure endpoint error: {data.get('error')}")

                parsed = data if isinstance(data, dict) else {}
                reply = _extract_reply_from_azure(parsed)
                finish_reason = _extract_finish_reason_from_azure(parsed)
                if not reply:
                    raise RuntimeError("Azure endpoint response missing expected choices/message content")

                collected_parts.append(reply)

                if not _looks_truncated(reply, finish_reason):
                    break
                if round_idx >= AZURE_AUTO_CONTINUE_ROUNDS:
                    break

                msgs.append({"role": "assistant", "content": reply})
                msgs.append(
                    {
                        "role": "user",
                        "content": "تابع من حيث توقفت في نفس الجواب، من آخر كلمة، دون إعادة المقدمة أو تكرار ما سبق.",
                    }
                )

            final_reply = "".join(collected_parts).strip()
            if final_reply:
                return final_reply

            raise RuntimeError("Azure endpoint returned empty reply")
    except httpx.HTTPStatusError as e:
        detail = ""
        try:
            detail = e.response.text[:500]
        except Exception:
            detail = ""
        raise RuntimeError(f"Azure HTTP {e.response.status_code}: {detail}")
    except httpx.TimeoutException:
        raise RuntimeError(f"Azure request timed out after {MODEL_HTTP_TIMEOUT_SECONDS} seconds")
    except Exception as e:
        raise RuntimeError(f"Azure endpoint unreachable: {e}")


async def _detect_and_translate(text: str) -> str:
    """Detect language of text and translate to Arabic if not already"""
    try:
        if not Translator:
            logger.warning("Translator class not imported")
            return text
        
        t = await _get_translator()
        if not t:
            logger.warning("Translator instance could not be initialized")
            return text
        
        # Detect the language
        detection = await t.detect(text)
        detected_lang = detection.get('lang', 'en') if isinstance(detection, dict) else getattr(detection, 'lang', 'en')
        
        logger.info("Detected language: %s", detected_lang)
        
        # If already in Arabic, return as-is
        if detected_lang in ('ar', 'und'):  # und = undetermined, likely Arabic
            logger.info("Response already in Arabic, not translating")
            return text
        
        # Otherwise translate to Arabic
        result = await t.translate(text, 'ar')
        
        # Result is a Translated object with .text attribute
        if hasattr(result, 'text'):
            return result.text
        else:
            logger.warning("Translation result has no .text attribute")
            return text
    except Exception as e:
        logger.error("Detection/translation failed: %s", str(e)[:100])
        return text


async def _stream_azure(history: List[dict]) -> AsyncGenerator[str, None]:
    reply = await _call_azure_endpoint(history)
    if not reply:
        raise RuntimeError("Azure endpoint unreachable or returned empty reply")
    
    logger.info("Raw Azure reply (first 300 chars): %s", reply[:300])
    
    # Detect language and translate to Arabic if needed
    arabic_reply = await _detect_and_translate(reply)
    
    logger.info("After language processing (first 300 chars): %s", arabic_reply[:300])
    
    # Stream character by character for better Arabic support
    for char in arabic_reply:
        yield char
        if char in (' ', '\n'):  # Small delay after spaces/newlines
            await asyncio.sleep(0.01)


def _ensure_model_file() -> bool:
    if os.path.exists(FINETUNED_MODEL_PATH):
        return True
    try:
        from huggingface_hub import hf_hub_download

        logger.info("Downloading fine-tuned GGUF from HF: %s/%s", FINETUNED_HF_REPO, FINETUNED_HF_FILE)
        model_dir = os.path.dirname(FINETUNED_MODEL_PATH)
        if model_dir:
            os.makedirs(model_dir, exist_ok=True)

        path = hf_hub_download(
            repo_id=FINETUNED_HF_REPO,
            filename=FINETUNED_HF_FILE,
            local_dir=model_dir or None,
        )

        if path != FINETUNED_MODEL_PATH and os.path.exists(path):
            os.replace(path, FINETUNED_MODEL_PATH)

        return os.path.exists(FINETUNED_MODEL_PATH)
    except Exception as e:
        logger.exception("Failed to download fine-tuned model: %s", e)
        return False


def _load_llm():
    global _llm_instance
    if _llm_instance is not None:
        return _llm_instance
    if not _ensure_model_file():
        return None
    try:
        from llama_cpp import Llama  # type: ignore[import]

        logger.info("Loading GGUF model from %s (ctx=%d, threads=%d)", FINETUNED_MODEL_PATH, FINETUNED_N_CTX, FINETUNED_N_THREADS)
        _llm_instance = Llama(
            model_path=FINETUNED_MODEL_PATH,
            n_ctx=FINETUNED_N_CTX,
            n_threads=FINETUNED_N_THREADS,
            chat_format="llama-3",
            verbose=False,
        )
        logger.info("GGUF model loaded")
        return _llm_instance
    except Exception as e:
        logger.exception("Failed to load GGUF model: %s", e)
        return None


def _run_finetuned_sync(history: List[dict]) -> str:
    llm = _load_llm()
    if llm is None:
        raise RuntimeError("Fine-tuned model not available")

    msgs = [{"role": "system", "content": SYSTEM_PROMPT}]
    for m in history:
        msgs.append({"role": m["role"], "content": m["content"]})

    out = llm.create_chat_completion(messages=msgs, temperature=0.4, max_tokens=256)
    return out["choices"][0]["message"]["content"].strip()


async def _call_finetuned(history: List[dict]) -> Optional[str]:
    try:
        async with _llm_lock:
            return await asyncio.to_thread(_run_finetuned_sync, history)
    except Exception as e:
        logger.warning("Fine-tuned model error, will fall back: %s", e)
        return None


async def _stream_finetuned(history: List[dict]) -> AsyncGenerator[str, None]:
    llm = _load_llm()
    if llm is None:
        raise RuntimeError("Fine-tuned model not available")

    msgs = [{"role": "system", "content": SYSTEM_PROMPT}]
    for m in history:
        msgs.append({"role": m["role"], "content": m["content"]})

    loop = asyncio.get_running_loop()
    q: asyncio.Queue = asyncio.Queue()
    sentinel = object()

    def producer() -> None:
        try:
            for chunk in llm.create_chat_completion(messages=msgs, temperature=0.4, max_tokens=256, stream=True):
                delta = chunk["choices"][0].get("delta", {}).get("content", "")
                if delta:
                    asyncio.run_coroutine_threadsafe(q.put(delta), loop)
        except Exception as e:
            asyncio.run_coroutine_threadsafe(q.put(e), loop)
        finally:
            asyncio.run_coroutine_threadsafe(q.put(sentinel), loop)

    async with _llm_lock:
        t = threading.Thread(target=producer, daemon=True)
        t.start()
        try:
            while True:
                item = await q.get()
                if item is sentinel:
                    return
                if isinstance(item, Exception):
                    raise item
                yield item
        finally:
            t.join(timeout=5)


async def _stream_local(history: List[dict]) -> AsyncGenerator[str, None]:
    reply = await _call_local_chatbot(history)
    if not reply:
        raise RuntimeError("Local chatbot unreachable or returned empty reply")
    for part in re.findall(r"\S+\s*", reply):
        yield part
        await asyncio.sleep(0.01)


async def _with_keepalive(gen: AsyncGenerator[str, None], interval_seconds: float) -> AsyncGenerator[str, None]:
    q: asyncio.Queue = asyncio.Queue()
    sentinel = object()

    async def producer() -> None:
        try:
            async for item in gen:
                await q.put(item)
        except Exception as e:
            await q.put(e)
        finally:
            await q.put(sentinel)

    task = asyncio.create_task(producer())
    try:
        while True:
            try:
                item = await asyncio.wait_for(q.get(), timeout=interval_seconds)
            except asyncio.TimeoutError:
                yield "event: ping\ndata: {}\n\n"
                continue

            if item is sentinel:
                break
            if isinstance(item, Exception):
                raise item

            yield item
    finally:
        if not task.done():
            task.cancel()
        with contextlib.suppress(asyncio.CancelledError):
            await task


@api.get("/")
async def root():
    return {"app": "Mizan", "status": "ok"}


@api.get("/health")
async def health():
    # Report the cache state without triggering the (slow) initial load —
    # startup warms the dataset in a background thread.
    dataset_records = _law_dataset_cache or []
    return {
        "status": "ok",
        "local_chatbot": bool(CHATBOT_LOCAL_URL),
        "llm": "configured" if GEMINI_API_KEY else "missing",
        "rag": {
            "enabled": bool(GEMINI_API_KEY),
            "configured": bool(GEMINI_API_KEY and dataset_records),
            "model": GEMINI_MODEL,
            "dataset_path": str(_resolve_law_dataset_path()),
            "dataset_loaded": bool(dataset_records),
            "dataset_loading": _law_dataset_cache is None,
            "dataset_rows": len(dataset_records),
            "top_k": RAG_TOP_K,
        },
        "finetuned": {
            "enabled": USE_FINETUNED,
            "model_path": FINETUNED_MODEL_PATH,
            "downloaded": os.path.exists(FINETUNED_MODEL_PATH),
            "loaded": _llm_instance is not None,
            "hf_repo": FINETUNED_HF_REPO,
        },
    }


@api.post("/auth/register", response_model=TokenResponse)
async def register(body: UserCreate):
    email = body.email.strip().lower()
    if not body.name.strip():
        raise HTTPException(400, "Name is required")
    if await db.users.find_one({"email": email}):
        raise HTTPException(409, "Email already registered")

    try:
        hashed_password = hash_password(body.password)
    except ValueError as exc:
        raise HTTPException(400, str(exc))

    user = User(
        name=body.name.strip(),
        email=email,
        hashed_password=hashed_password,
    )
    await db.users.insert_one(user.model_dump())
    token = create_access_token(user.id)
    return TokenResponse(token=token)


@api.post("/auth/login", response_model=TokenResponse)
async def login(body: UserLogin):
    email = body.email.strip().lower()
    user_doc = await db.users.find_one({"email": email}, {"_id": 0})
    if not user_doc or not verify_password(body.password, user_doc["hashed_password"]):
        raise HTTPException(401, "Invalid email or password")

    token = create_access_token(user_doc["id"])
    return TokenResponse(token=token)


@api.post("/auth/guest", response_model=TokenResponse)
async def guest_login():
    guest_user = await _create_guest_user()
    token = create_access_token(guest_user.id)
    return TokenResponse(token=token)


@api.get("/auth/me", response_model=UserOut)
async def get_me(current_user: UserOut = Depends(get_current_user)):
    return current_user


@api.post("/debug/raw")
async def debug_raw(request: Request):
    """Temporary debug endpoint: logs raw request body for troubleshooting malformed requests."""
    try:
        body = await request.body()
        try:
            text = body.decode("utf-8")
        except Exception:
            text = repr(body)
        client_host = getattr(request.client, "host", None) if request.client else None
        logger.info("DEBUG RAW received from=%s path=%s body=%s", client_host, request.url.path, (text[:10000] if isinstance(text, str) else str(text)))
        return {"received": text, "length": len(body)}
    except Exception as e:
        logger.exception("Failed to read raw body: %s", e)
        raise HTTPException(500, "Failed to read body")


@api.post("/debug/azure_test")
async def debug_azure_test():
    """Temporary debug endpoint: call AZURE endpoint with minimal payload and return status/result."""
    if not AZURE_ML_ENDPOINT or not AZURE_ML_API_KEY:
        raise HTTPException(400, "Azure endpoint or API key not configured")
    history = [{"role": "user", "content": "سلام"}]
    try:
        result = await _call_azure_endpoint(history)
        return {"ok": True, "reply_preview": (result[:400] if result else None)}
    except Exception as e:
        logger.exception("Azure test call failed: %s", e)
        raise HTTPException(502, str(e))


@api.post("/chats", response_model=Chat)
async def create_chat(body: ChatCreate, current_user: UserOut = Depends(get_current_user)):
    logger.info("POST /api/chats title=%s user=%s", body.title or "<default>", current_user.id)
    chat_obj = Chat(title=body.title or "محادثة جديدة")
    chat_data = chat_obj.model_dump()
    chat_data["user_id"] = current_user.id
    await db.chats.insert_one(chat_data)
    return chat_obj


@api.get("/chats", response_model=List[Chat])
async def list_chats(current_user: UserOut = Depends(get_current_user)):
    docs = await db.chats.find({"user_id": current_user.id}, {"_id": 0}).sort("updated_at", -1).to_list(500)
    return docs


@api.get("/chats/{chat_id}", response_model=Chat)
async def get_chat(chat_id: str, current_user: UserOut = Depends(get_current_user)):
    doc = await db.chats.find_one({"id": chat_id, "user_id": current_user.id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Chat not found")
    return doc


@api.delete("/chats/{chat_id}")
async def delete_chat(chat_id: str, current_user: UserOut = Depends(get_current_user)):
    chat_doc = await db.chats.find_one({"id": chat_id, "user_id": current_user.id}, {"_id": 0})
    if not chat_doc:
        raise HTTPException(404, "Chat not found")
    await db.messages.delete_many({"chat_id": chat_id})
    res = await db.chats.delete_one({"id": chat_id, "user_id": current_user.id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Chat not found")
    return {"ok": True}


@api.patch("/chats/{chat_id}", response_model=Chat)
async def update_chat(chat_id: str, body: ChatUpdate, current_user: UserOut = Depends(get_current_user)):
    doc = await db.chats.find_one({"id": chat_id, "user_id": current_user.id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Chat not found")
    title = body.title.strip() or "محادثة جديدة"
    updated_at = datetime.now(timezone.utc).isoformat()
    await db.chats.update_one(
        {"id": chat_id, "user_id": current_user.id},
        {"$set": {"title": title, "updated_at": updated_at}},
    )
    doc["title"] = title
    doc["updated_at"] = updated_at
    return doc


@api.get("/chats/{chat_id}/messages", response_model=List[Message])
async def list_messages(chat_id: str, current_user: UserOut = Depends(get_current_user)):
    chat_doc = await db.chats.find_one({"id": chat_id, "user_id": current_user.id}, {"_id": 0})
    if not chat_doc:
        raise HTTPException(404, "Chat not found")
    docs = await db.messages.find({"chat_id": chat_id}, {"_id": 0}).sort("created_at", 1).to_list(1000)
    return docs


@api.post("/chat", response_model=Message)
async def send_message(body: SendMessageBody, current_user: UserOut = Depends(get_current_user)):
    logger.info("POST /api/chat chat_id=%s user=%s use_local=%s use_azure=%s use_finetuned=%s", body.chat_id, current_user.id, body.use_local, body.use_azure, body.use_finetuned)
    chat = await db.chats.find_one({"id": body.chat_id, "user_id": current_user.id}, {"_id": 0})
    if not chat:
        raise HTTPException(404, "Chat not found")
    if not body.content.strip():
        raise HTTPException(400, "Empty message")

    user_msg = Message(chat_id=body.chat_id, role="user", content=body.content.strip())
    await db.messages.insert_one(user_msg.model_dump())
    _log_message_event(body.chat_id, current_user.id, "user", "client", body.content.strip())

    history_docs = await db.messages.find({"chat_id": body.chat_id}, {"_id": 0}).sort("created_at", 1).to_list(1000)
    history = [{"role": d["role"], "content": d["content"]} for d in history_docs]

    pdf_requested = _wants_pdf(body.content)
    model_history = _append_pdf_note_to_history(history) if pdf_requested else history

    reply_text: Optional[str] = None
    source = "unavailable"
    gemini_error: Optional[str] = None

    short_reply = get_short_circuit_response(body.content.strip())
    if short_reply is not None:
        reply_text = short_reply
        source = "short_circuit"
        logger.info("Reply selected from static response (short_circuit) chat_id=%s user=%s", body.chat_id, current_user.id)
    elif body.use_local:
        reply_text = await _call_local_chatbot(model_history)
        if reply_text is not None:
            source = "local_url"
            logger.info("Reply selected from local chatbot URL chat_id=%s user=%s", body.chat_id, current_user.id)

    if reply_text is None and GEMINI_API_KEY:
        try:
            reply_text = await _call_gemini_rag(model_history)
            if reply_text is not None:
                source = "gemini_rag"
                logger.info("Reply selected from Gemini RAG chat_id=%s user=%s", body.chat_id, current_user.id)
            else:
                gemini_error = "Gemini API is not configured"
        except Exception as e:
            gemini_error = str(e)
            logger.warning("Gemini call failed: %s", gemini_error)

    if reply_text is None and (body.use_finetuned or USE_FINETUNED):
        reply_text = await _call_finetuned(history)
        if reply_text is not None:
            source = "finetuned"
            logger.info("Reply selected from finetuned model chat_id=%s user=%s", body.chat_id, current_user.id)

    if reply_text is None:
        raise HTTPException(
            503,
            gemini_error or "No available model response. Ensure Gemini is configured or enable local/fine-tuned model.",
        )

    _log_answer_source(source, body.chat_id, current_user.id, reply_text)
    ai_msg = Message(chat_id=body.chat_id, role="assistant", content=reply_text, source=source)
    if pdf_requested and source != "short_circuit":
        await _attach_answer_pdf(ai_msg, chat.get("title", ""))
    await db.messages.insert_one(ai_msg.model_dump())
    _log_message_event(body.chat_id, current_user.id, "assistant", "static" if source == "short_circuit" else "model", reply_text)

    new_title = chat.get("title", "محادثة جديدة")
    if new_title == "محادثة جديدة":
        new_title = await _generate_title_from_first_message(body.content)
    await db.chats.update_one({"id": body.chat_id}, {"$set": {"title": new_title, "updated_at": datetime.now(timezone.utc).isoformat()}})

    return ai_msg


@api.post("/documents/analyze", response_model=Message)
async def analyze_document(
    chat_id: str = Form(...),
    question: str = Form(""),
    file: UploadFile = File(...),
    current_user: UserOut = Depends(get_current_user),
):
    chat = await db.chats.find_one({"id": chat_id, "user_id": current_user.id}, {"_id": 0})
    if not chat:
        raise HTTPException(404, "Chat not found")
    if not GEMINI_API_KEY:
        raise HTTPException(503, "Document analysis requires GEMINI_API_KEY to be configured")

    mime_type = file.content_type or "application/octet-stream"
    if mime_type not in DOCUMENT_ALLOWED_MIME_TYPES:
        raise HTTPException(400, f"Unsupported file type: {mime_type}. Allowed: {', '.join(sorted(DOCUMENT_ALLOWED_MIME_TYPES))}")

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(400, "Uploaded file is empty")
    if len(file_bytes) > DOCUMENT_MAX_BYTES:
        raise HTTPException(400, f"File too large ({len(file_bytes)} bytes). Max is {DOCUMENT_MAX_BYTES} bytes")

    user_message_text = question.strip() or f"[Uploaded document: {file.filename}]"
    original_filename = os.path.basename(file.filename or "document")
    user_msg = Message(
        chat_id=chat_id,
        role="user",
        content=user_message_text,
        attachment=Attachment(filename=original_filename, mime_type=mime_type, size=len(file_bytes)),
    )
    _save_uploaded_document(chat_id, user_msg.id, original_filename, file_bytes)
    await db.messages.insert_one(user_msg.model_dump())
    _log_message_event(chat_id, current_user.id, "user", "document_upload", user_message_text)

    try:
        reply_text = await _call_gemini_document(question, file_bytes, mime_type)
    except Exception as e:
        logger.warning("Document analysis failed: %s", e)
        raise HTTPException(502, str(e))

    if not reply_text:
        raise HTTPException(502, "Gemini returned no content for this document")

    source = "gemini_document"
    _log_answer_source(source, chat_id, current_user.id, reply_text)
    ai_msg = Message(chat_id=chat_id, role="assistant", content=reply_text, source=source)
    await db.messages.insert_one(ai_msg.model_dump())
    _log_message_event(chat_id, current_user.id, "assistant", "model", reply_text)

    new_title = chat.get("title", "محادثة جديدة")
    if new_title == "محادثة جديدة":
        new_title = await _generate_title_from_first_message(user_message_text)
    await db.chats.update_one({"id": chat_id}, {"$set": {"title": new_title, "updated_at": datetime.now(timezone.utc).isoformat()}})

    return ai_msg


@api.get("/documents/{message_id}")
async def get_document(message_id: str, current_user: UserOut = Depends(get_current_user)):
    message_doc = await db.messages.find_one({"id": message_id}, {"_id": 0})
    if not message_doc or not message_doc.get("attachment"):
        raise HTTPException(404, "Document not found")

    chat_doc = await db.chats.find_one({"id": message_doc["chat_id"], "user_id": current_user.id}, {"_id": 0})
    if not chat_doc:
        raise HTTPException(404, "Document not found")

    file_path = _resolve_uploaded_document(message_doc["chat_id"], message_id)
    if not file_path:
        raise HTTPException(404, "Document file is no longer available")

    attachment = message_doc["attachment"]
    return FileResponse(
        path=file_path,
        media_type=attachment.get("mime_type", "application/octet-stream"),
        filename=attachment.get("filename", file_path.name),
    )


@api.post("/chat/stream")
async def send_message_stream(body: SendMessageBody, current_user: UserOut = Depends(get_current_user)):
    logger.info("POST /api/chat/stream chat_id=%s user=%s use_local=%s use_azure=%s use_finetuned=%s", body.chat_id, current_user.id, body.use_local, body.use_azure, body.use_finetuned)
    chat = await db.chats.find_one({"id": body.chat_id, "user_id": current_user.id}, {"_id": 0})
    if not chat:
        raise HTTPException(404, "Chat not found")
    if not body.content.strip():
        raise HTTPException(400, "Empty message")

    user_msg = Message(chat_id=body.chat_id, role="user", content=body.content.strip())
    await db.messages.insert_one(user_msg.model_dump())
    _log_message_event(body.chat_id, current_user.id, "user", "client", body.content.strip())

    history_docs = await db.messages.find({"chat_id": body.chat_id}, {"_id": 0}).sort("created_at", 1).to_list(1000)
    history = [{"role": d["role"], "content": d["content"]} for d in history_docs]
    logger.info("History for chat %s: found %d messages", body.chat_id, len(history_docs))

    pdf_requested = _wants_pdf(body.content)
    model_history = _append_pdf_note_to_history(history) if pdf_requested else history

    use_local_now = body.use_local and bool(CHATBOT_LOCAL_URL)
    use_gemini_now = bool(GEMINI_API_KEY) and not use_local_now
    use_finetuned_now = (body.use_finetuned or USE_FINETUNED) and not use_local_now and not use_gemini_now

    short_reply = get_short_circuit_response(body.content.strip())
    if short_reply is not None:
        async def event_gen():
            source = "short_circuit"
            logger.info("Streaming reply selected from static response (short_circuit) chat_id=%s user=%s", body.chat_id, current_user.id)
            _log_answer_source(source, body.chat_id, current_user.id, short_reply)
            ai_msg = Message(chat_id=body.chat_id, role="assistant", content=short_reply, source=source)
            await db.messages.insert_one(ai_msg.model_dump())
            _log_message_event(body.chat_id, current_user.id, "assistant", "static", short_reply)

            new_title = chat.get("title", "محادثة جديدة")
            if new_title == "محادثة جديدة":
                new_title = await _generate_title_from_first_message(body.content)
            await db.chats.update_one({"id": body.chat_id}, {"$set": {"title": new_title, "updated_at": datetime.now(timezone.utc).isoformat()}})

            yield f"event: token\ndata: {json.dumps({'text': short_reply}, ensure_ascii=False)}\n\n"
            yield f"event: done\ndata: {json.dumps({'message_id': ai_msg.id, 'source': source}, ensure_ascii=False)}\n\n"

        return StreamingResponse(
            event_gen(),
            media_type="text/event-stream",
            headers={"Cache-Control": "no-cache", "Connection": "keep-alive", "X-Accel-Buffering": "no"},
        )

    async def event_gen():
        full_chunks: List[str] = []
        source = "local_url" if use_local_now else ("gemini_rag" if use_gemini_now else "finetuned")
        logger.info("Streaming reply selected from %s chat_id=%s user=%s", source, body.chat_id, current_user.id)
        try:
            if use_local_now:
                gen = _stream_local(model_history)
            elif use_gemini_now:
                gen = _stream_gemini_rag(model_history)
            elif use_finetuned_now:
                gen = _stream_finetuned(model_history)
            else:
                raise HTTPException(
                    503,
                    "No available model response. Ensure Gemini is configured or enable local/fine-tuned model.",
                )

            async for token in _with_keepalive(gen, STREAM_KEEPALIVE_SECONDS):
                if token.startswith("event: ping"):
                    yield token
                    continue
                full_chunks.append(token)
                yield f"event: token\ndata: {json.dumps({'text': token}, ensure_ascii=False)}\n\n"

            reply_text = "".join(full_chunks)
            _log_answer_source(source, body.chat_id, current_user.id, reply_text)
            ai_msg = Message(chat_id=body.chat_id, role="assistant", content=reply_text, source=source)
            if pdf_requested:
                await _attach_answer_pdf(ai_msg, chat.get("title", ""))
            await db.messages.insert_one(ai_msg.model_dump())
            _log_message_event(body.chat_id, current_user.id, "assistant", "model", reply_text)

            new_title = chat.get("title", "محادثة جديدة")
            if new_title == "محادثة جديدة":
                new_title = await _generate_title_from_first_message(body.content)
            await db.chats.update_one({"id": body.chat_id}, {"$set": {"title": new_title, "updated_at": datetime.now(timezone.utc).isoformat()}})

            yield f"event: done\ndata: {json.dumps({'message_id': ai_msg.id, 'source': source}, ensure_ascii=False)}\n\n"
        except Exception as e:
            logger.exception("Stream error")
            yield f"event: error\ndata: {json.dumps({'error': str(e)}, ensure_ascii=False)}\n\n"

    return StreamingResponse(
        event_gen(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "Connection": "keep-alive", "X-Accel-Buffering": "no"},
    )


app.include_router(api)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","),
    # Optional regex for origins that vary per deployment (e.g. Vercel preview
    # URLs like https://mizan-abc123-team.vercel.app). Set CORS_ORIGIN_REGEX
    # alongside CORS_ORIGINS; either match allows the request.
    allow_origin_regex=os.environ.get("CORS_ORIGIN_REGEX") or None,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
async def warm_law_dataset():
    # Download/convert/index the law dataset off the event loop so the server
    # answers requests immediately while the RAG index warms up in background.
    threading.Thread(target=_load_law_dataset, daemon=True).start()


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()


